import numpy as np
import pandas as pd
import os
from scipy.stats import skew, kurtosis

def calc_iqr(series):
    q1 = series.quantile(.25)
    q3 = series.quantile(.75)
    iqr = q3 - q1
    outliers = ((series < (q1 - 1.5* iqr)) | (series > (q3 + 1.5*iqr))).sum()
    return iqr, outliers

def get_summary_df(df):
    grouped = df.groupby('SUBJECT')
    results = []
    for subject, group in grouped:
        subject_stats = {'SUBJECT' : subject}
        for time in ['T1', 'T2', 'T1+T2', 'T3']:
            series = group[time].dropna()
            series = series[series > 150]
            if len(series) > 0:
                subject_stats[f'{time}_mean'] = series.mean()
                subject_stats[f'{time}_median'] = series.median()
                subject_stats[f'{time}_std'] = series.std()
                subject_stats[f'{time}_cv'] = series.std()/ series.mean()
                subject_stats[f'{time}_skewness'] = skew(series)
                subject_stats[f'{time}_kurtosis'] = kurtosis(series)
                iqr, outlier = calc_iqr(series)
                subject_stats[f'{time}_iqr'] = iqr
                subject_stats[f'{time}_outlier'] = outlier
                subject_stats[f'{time}_nsamples'] = len(series)
            else:
                subject_stats[f"{time}_mean"] = np.nan
                subject_stats[f"{time}_median"] = np.nan
                subject_stats[f"{time}_std"] = np.nan
                subject_stats[f'{time}_cv'] = np.nan
                subject_stats[f"{time}_skewness"] = np.nan
                subject_stats[f"{time}_kurtosis"] = np.nan
                subject_stats[f"{time}_IQR"] = np.nan
                subject_stats[f"{time}_outliers"] = np.nan
                subject_stats[f'{time}_nsamples'] = 0
        results.append(subject_stats)
    overall_stats = {'SUBJECT': 'ALL'}
    for time in ['T1', 'T2', 'T1+T2', 'T3']:
        series = df[time].dropna()
        series = series[series > 150]
        if len(series) > 0:
            overall_stats[f'{time}_mean'] = series.mean()
            overall_stats[f'{time}_median'] = series.median()
            overall_stats[f'{time}_std'] = series.std()
            overall_stats[f'{time}_cv'] = series.std()/ series.mean()
            overall_stats[f'{time}_skewness'] = skew(series)
            overall_stats[f'{time}_kurtosis'] = kurtosis(series)
            iqr, outlier = calc_iqr(series)
            overall_stats[f'{time}_iqr'] = iqr
            overall_stats[f'{time}_outlier'] = outlier
            overall_stats[f'{time}_nsamples'] = len(series)
        else:
            overall_stats[f"{time}_mean"] = np.nan
            overall_stats[f"{time}_median"] = np.nan
            overall_stats[f"{time}_std"] = np.nan
            overall_stats[f'{time}_cv'] = np.nan
            overall_stats[f"{time}_skewness"] = np.nan
            overall_stats[f"{time}_kurtosis"] = np.nan
            overall_stats[f"{time}_iqr"] = np.nan
            overall_stats[f"{time}_outlier"] = np.nan
            overall_stats[f'{time}_nsamples'] = 0

    # Append overall statistics to results
    results.append(overall_stats)
    return pd.DataFrame(results)

def main():
    path = 'stat/subject_tables_Stop'
    info_file = pd.read_excel('stat/Status Reclutamento.xlsx', usecols=  'B,D,G', header = None)
    info_file = info_file.iloc[11:, :]
    info_file.columns = ['USER', 'TYPE', 'LABELED']
    info_file = info_file.loc[info_file['LABELED'] == 'Y', ['USER', 'TYPE']]
    stop_list = info_file[info_file['TYPE'] == 'Stop']['USER'].to_list()
    stop_list = [f'IDU{int(stop):03d}' for stop in stop_list]
    file_list = os.listdir(path)
    dfs_stop = []
    for file in file_list:
        df = pd.read_csv(os.path.join(path,file))
        if file[:6] in stop_list:
            dfs_stop.append(df)
    dfs_stop = pd.concat(dfs_stop, ignore_index=True)
    dfs_stop = dfs_stop[dfs_stop['Previous Go - Video'] > 0]
    dfs_stop = dfs_stop[dfs_stop['Previous Go TOT'] > 0]
    dfs_stop = dfs_stop[dfs_stop['T1'] > 150]
    results = get_summary_df(dfs_stop)
    results.to_csv('stat/summary_times_stop_total.csv', index = False)
    # print(results)
    median_stimuli = dfs_stop['Previous Go - Video'].median()
    print(median_stimuli)
    group_1 = dfs_stop[dfs_stop['Previous Go - Video'] <= median_stimuli]
    group_2 = dfs_stop[dfs_stop['Previous Go - Video'] > median_stimuli]
    results = get_summary_df(group_1)
    results.to_csv('stat/summary_times_stop_under_median.csv', index = False)
    # print(results)
    results = get_summary_df(group_2)
    results.to_csv('stat/summary_times_stop_over_median.csv', index = False)
    # print(results)
    print(group_1)
    print(group_2)
    # for time in ['T1', 'T2', 'T3']:

def kendal_correlation_big5():
    from scipy.stats import kendalltau

    times_path = 'stat/summary_times_stop_total.csv'
    big5_path =  r'C:\Users\marco\OneDrive - unifi.it\Cartella Ricerca Progetto Destini Condisa\Sperimentazione IMAGINE\BIG-5_values.csv'
    times_df = pd.read_csv(times_path)
    big5_df = pd.read_csv(big5_path)
    # Define the columns of interest
    # times_columns = [
    #     'T1_mean', 'T1_kurtosis','T1_cv', 
    #     'T2_mean', 'T2_kurtosis','T2_cv', 
    #     'T3_mean',  'T3_kurtosis','T3_cv',
    #     'T1+T2_mean', 'T1+T2_kurtosis' ,'T1+T2_cv',
    # ]

    big5_columns = ["Extraversion","Agreeableness","Conscientiousness","Neuroticism","Openness"]

    # Ensure the SUBJECT column is used to align both dataframes
    times_df = times_df.set_index('SUBJECT')
    big5_df = big5_df.set_index('subject')

    # Ensure alignment of subjects between both dataframes
    common_subjects = times_df.index.intersection(big5_df.index)
    common_subjects = common_subjects.drop('IDU040')
    times_columns = times_df.columns
    times_df = times_df.loc[common_subjects, times_columns]
    # times_columns = [meas for meas in times_columns if 'n_samp' not in meas]
    big5_df = big5_df.loc[common_subjects, big5_columns]
    # Prepare a results dictionary
    results = []
    # Compute correlations
    for t_col in times_columns:
        # if 'samp' in t_col:
        #     continue
        for b_col in big5_columns:
            # if 'samp' in b_col:
            #     continue
            corr, pval = kendalltau(times_df[t_col], big5_df[b_col])
            results.append({'Times Metric': t_col, 'big5 Trait': b_col, 'Correlation': corr, 'P-value': pval})

    # Format the results into a DataFrame
    results_df = pd.DataFrame(results)

    # Sort by p-value for easier interpretation
    meas = 'big5'
    results_df.to_csv(f'stat/correlation_results_{meas}.csv', index=False)
    results_df = results_df.sort_values(by='P-value')
    results_df.to_csv(f'stat/correlation_results_{meas}_ordered.csv', index=False)

    return results_df

def kendal_correlation_bis11():
    from scipy.stats import kendalltau

    times_path = 'stat/summary_times_stop_total.csv'
    bis11_path =  r'C:\Users\marco\OneDrive - unifi.it\Cartella Ricerca Progetto Destini Condisa\Sperimentazione IMAGINE\BIS-11_values.csv'
    times_df = pd.read_csv(times_path)
    bis11_df = pd.read_csv(bis11_path)

    bis11_columns = ["Attenzione","Comp_motorio","Autocontrollo","Compl_cogn","Perseveranza","Instabilita_cog","Impulsivita_cog","Impulsivita_mot","Impulsivita_non_pian"]

    # Ensure the SUBJECT column is used to align both dataframes
    times_df = times_df.set_index('SUBJECT')
    bis11_df = bis11_df.set_index('subject')

    # Ensure alignment of subjects between both dataframes
    common_subjects = times_df.index.intersection(bis11_df.index)
    common_subjects = common_subjects.drop('IDU040')
    times_columns = times_df.columns
    times_columns = [meas for meas in times_columns if 'n_samp' not in meas]
    times_df = times_df.loc[common_subjects, times_columns]
    bis11_df = bis11_df.loc[common_subjects, bis11_columns]
    # Prepare a results dictionary
    results = []
    # Compute correlations
    for t_col in times_columns:
        # if 'samp' in t_col:
        #     continue
        for b_col in bis11_columns:
            # if 'samp' in b_col:
            #     continue
            corr, pval = kendalltau(times_df[t_col], bis11_df[b_col])
            results.append({'Times Metric': t_col, 'bis11 Trait': b_col, 'Correlation': corr, 'P-value': pval})

    # Format the results into a DataFrame
    results_df = pd.DataFrame(results)

    # Sort by p-value for easier interpretation
    meas = 'bis11'
    results_df.to_csv(f'stat/correlation_results_{meas}.csv', index=False)
    results_df = results_df.sort_values(by='P-value')
    results_df.to_csv(f'stat/correlation_results_{meas}_ordered.csv', index=False)
    return results_df

def kendal_correlation(df):
    from scipy.stats import kendalltau
    meas = df.split('/')[1].split('.')[0]

    times_path = 'stat/summary_times_stop_total.csv'
    times_df = pd.read_csv(times_path)
    meas_df = pd.read_csv(df)
    print(meas_df)
    times_df = times_df.set_index('SUBJECT')
    meas_df = meas_df.set_index('SUBJECT')
    common_subjects = times_df.index.intersection(meas_df.index)
    common_subjects = common_subjects.drop('IDU040')
    common_subjects = common_subjects.drop('ALL')
    times_columns = times_df.columns
    times_df = times_df.loc[common_subjects, times_columns]
    meas_columns = meas_df.columns
    meas_df = meas_df.loc[common_subjects, meas_columns]
    # Prepare a results dictionary
    results = []
    print(times_columns)
    print(meas_columns)
    # Compute correlations
    for t_col in times_columns:
        # if 'samp' in t_col:
        #     continue
        for b_col in meas_columns:
            # if 'samp' in b_col:
            #     continue
            # Pearson correlation and p-value
            corr, pval = kendalltau(times_df[t_col], meas_df[b_col])
            results.append({'Times Metric': t_col, f'{meas} Trait': b_col, 'Correlation': corr, 'P-value': pval})
    results_df = pd.DataFrame(results)

    # Sort by p-value for easier interpretation
    # results_df = results_df.sort_values(by='P-value')
    results_df.to_csv(f'stat/correlation_results_{meas}.csv', index=False)
    results_df = results_df.sort_values(by='P-value')
    results_df.to_csv(f'stat/correlation_results_{meas}_ordered.csv', index=False)
    return results_df
# Example usage
def plot_table( test = 'bis11'):
    import matplotlib.pyplot as plt
    from matplotlib.table import Table
   
    file = f'stat/correlation_results_{test}.csv'
    data = pd.read_csv(file)
    data_col = data[f'{test} Trait'].unique()

    data["Correlation (P-value)"] = data.apply(
        lambda row: f"{row['Correlation']:.2f} ({row['P-value']:.2f})", axis=1
    )

    times_path = 'stat/summary_times_stop_total.csv'
    times_df = pd.read_csv(times_path)
    times_df = times_df.set_index('SUBJECT')
    ordered_metrics = times_df.columns
    # ordered_metrics = [
    # "T1_mean", "T1_kurtosis", "T1_cv",
    # "T2_mean", "T2_kurtosis", "T2_cv",
    # "T3_mean", "T3_kurtosis", "T3_cv",
    # "T1+T2_mean", "T1+T2_kurtosis", "T1+T2_cv"
    # ]
    # data["Times Metric"] = pd.Categorical(data["Times Metric"], categories=ordered_metrics, ordered=True)
    data["Times Metric"] = pd.Categorical(data["Times Metric"], categories=ordered_metrics, ordered=False)
    data = data.sort_values("Times Metric")
    table_data = data.pivot(
        index="Times Metric", columns=f"{test} Trait", values="Correlation (P-value)"
    ).reindex(columns=data_col)


    p_values = data.pivot(
        index="Times Metric", columns=f"{test} Trait", values="P-value"
    ).reindex(columns=data_col)

    # Create a matplotlib figure for the table
    fig, ax = plt.subplots(figsize=(20, 12))
    ax.axis("tight")
    ax.axis("off")

    # Create the table
    table = ax.table(
        cellText=table_data.values,
        rowLabels=table_data.index,
        colLabels=table_data.columns,
        loc="center",
        cellLoc="center",
        colLoc="center",
    )

    for i, row_key in enumerate(table_data.index):
        for j, col_key in enumerate(table_data.columns):
            p_value = p_values.loc[row_key, col_key]
            cell = table[(i + 1, j)]  # Offset for headers
            if p_value < 0.05:
                cell.set_text_props(weight="bold", color='red')
    # Add a title
    plt.title(f"Correlation (P-value) Table {test}", fontsize=14, weight="bold")

    # Show the plot
    plt.savefig(f"stat/correlation_table_{test}.png")
    
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Correlation Table"

    # Add the header row
    headers = ["Time Metric"] + list(table_data.columns)
    ws.append(headers)

    # Add the data rows
    for index, row in table_data.iterrows():
        ws.append([index] + row.tolist())

    # Style the table
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column)):
        time_metric = list(table_data.index)[i]
        for j, cell in enumerate(row):
            col_key = table_data.columns[j]
            p_value = p_values.loc[time_metric, col_key]

            cell.alignment = Alignment(horizontal="center", vertical="center")
            if p_value < 0.05:
                cell.font = Font(bold=True)

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Save the Excel file
    output_path = f"stat/correlation_table_{test}.xlsx"  # Update this to your desired output path
    wb.save(output_path)
    print(f"Table successfully saved to {output_path}")

def get_significative_correlation():
    import re
    import time
    path = 'stat'
    files = [os.path.join(path, file) for file in os.listdir(path) if 'correlation_table' in file and '.xlsx' in file and file != 'correlation_table.xlsx']
    with open('stat/significative_results.txt', 'w') as f:
        for file in files:
            print(file)
            df = pd.read_excel(file)
            # df.set_index('Time Metric')
            for idx, row in df.iterrows():
                time_metric = row['Time Metric']  # Access the value in the 'Time Metric' column
                for column in df.columns:
                    if column == 'Time Metric':
                        continue  # Skip the 'Time Metric' column
                    match = re.match(r'([-\d.]+) \(([\d.]+)\)', str(row[column]))
                    # print(row[column])
                    if match:
                        value = float(match.group(1))
                        p_value = float(match.group(2))
                        if p_value <= 0.05:
                            print(time_metric,column)
                            # time.sleep(0.5)
                            f.write(f"{time_metric}/{column} {value} ({p_value})\n")

if __name__ == '__main__':
    # test = 'stroop_incongruent'
    # kendal_correlation(f'stat/{test}.csv')
    # plot_table(f'{test}')
    kendal_correlation_bis11()
    plot_table(test='bis11')
    print(1)    
    kendal_correlation_big5()
    plot_table(test='big5')
    print(2)
    test = 'stroop_all'
    kendal_correlation(f'stat/{test}.csv')
    print(3)
    plot_table(f'{test}')
    test = 'stroop_congruent'
    kendal_correlation(f'stat/{test}.csv')
    print(3)
    plot_table(f'{test}')
    test = 'stroop_incongruent'
    kendal_correlation(f'stat/{test}.csv')
    print(3)
    plot_table(f'{test}')
    # kendal_correlation_big5()
    # plot_table( test = 'big5')
    # kendal_correlation_bis11()
    # plot_table( test = 'bis11')
    get_significative_correlation()
